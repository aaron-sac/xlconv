#! /usr/bin/env python3

import sys
import io
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import DEFAULT_FONT
import tempfile
import csv

MAGIC = "~/~/~/SHEET_START: %s"

def encode(source, target):
    input = io.open(source, 'rb')
    output = io.open(target, 'w')
    '''Convert from excel into a weird multi sheet supporting CSV text format'''
    workbook_file = tempfile.TemporaryFile(mode='w+b')
    workbook_file.write(input.read())
    workbook = openpyxl.load_workbook(workbook_file, read_only=True)
    csv_writer = csv.writer(output, delimiter=',', quotechar='"', quoting=csv.QUOTE_NONNUMERIC, lineterminator='\n')
    for sheet_name in workbook.sheetnames:
        output.write(MAGIC % sheet_name + '\n')
        for row in workbook[sheet_name]:
            cells = [c.value.replace('\r\n', '\n').replace('\r', '\n') if isinstance(c.value, str) else c.value for c in row]
            if any(cells):
                csv_writer.writerow(cells)

def decode(source, target):
    input = io.open(source, 'r')
    output = io.open(target, 'wb')
    '''Convert from a weird multi sheet supporting CSV format to excel'''
    DEFAULT_FONT.name = 'Consolas'
    workbook = openpyxl.Workbook()
    for sheet_name in workbook.sheetnames:
        workbook.remove(workbook[sheet_name])
    sheets = []
    sheet_file = None
    for line in input.readlines():
        if line.startswith(MAGIC % ''):
            sheet_file = tempfile.TemporaryFile(mode='r+')
            sheets.append((line.replace(MAGIC % '', '').strip(), sheet_file))
            continue
        elif line.replace(',', '').strip() == "":
            continue
        line = line.replace(',True', ',"TRUE"').replace(',False', ',"FALSE"')
        sheet_file.write(line)
    for (sheet_name, sheet_file) in sheets:
        sheet_file.seek(0)
        workbook.create_sheet(sheet_name)
        sheet = workbook[sheet_name]
        reader = csv.reader(sheet_file, delimiter=',', quotechar='"', quoting=csv.QUOTE_NONNUMERIC)
        for row in reader:
            while row and (str(row[-1]).strip() == "" or row[-1] is None):
                row.pop()
            if len(row) == 0:
                continue
            sheet.append([v if v != "" else None for v in row])
        if sheet_name != 'meta':
            sheet.freeze_panes = 'A2'
            header = openpyxl.styles.Font(bold=True, name="Consolas")
            # Enumerate the cells in the second row
            for cell in sheet["1:1"]:
                cell.font = header
        for column_cells in sheet.columns:
            new_column_length = max(len(str(cell.value)) for cell in column_cells)
            new_column_letter = (get_column_letter(column_cells[0].column))
            new_column_length = min(new_column_length, 50)
            if new_column_length > 0:
                sheet.column_dimensions[new_column_letter].width = new_column_length*1.23
    
    workbook.save(output)


def argument_exists(args):
    for a in args:
        if a in sys.argv:
            return True
    return False

def get_argument_value(args, default):
    for a in args:
        if sys.argv.index(a) < len(sys.argv) - 1:
            return sys.argv[sys.argv.index(a) + 1]
    return default


def passthrough(source, target):
    tmp_file = tempfile.NamedTemporaryFile(mode='r+')
    encode(source, tmp_file.name)
    decode(tmp_file.name, target)

def main():
    fn = None
    if argument_exists(['-e', '--encode']):
        fn = encode
    elif argument_exists(['-d', '--decode']):
        fn = decode
    elif argument_exists(['-p', '--passthrough']):
        fn = passthrough

    source = get_argument_value(['-s', '--source'], sys.stdin.fileno())
    target = get_argument_value(['-t', '--target'], sys.stdin.fileno())

    if fn is not None:
        fn(source, target)